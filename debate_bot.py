import logging, os, json, asyncio, sys
from datetime import datetime
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove, ChatMember
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler
import openpyxl
from openpyxl import load_workbook

BOT_TOKEN     = "8688745859:AAFcShsQng2sBVCruJ1c5yplyzwN9iuXn-g"
ADMIN_IDS     = [7968570881]
TG_CHANNEL    = "https://t.me/centraldebate_TIM"
TG_CHANNEL_ID = "@centraldebate_TIM"
IG_PAGE       = "soon"
DATA_FILE     = "registrations.xlsx"
EVENT_FILE    = "event_data.json"
USERS_FILE    = "all_users.json"

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)

(LANG, MAIN_MENU, SU_NAME, SU_SCHOOL, SU_CLASS, SU_AGE, SU_ENGLISH,
 SU_FOLLOW, SU_CONFIRM, SU_EDIT_MENU, SU_EDIT_FIELD,
 BC_WAIT, SD_WAIT, ST_WAIT, SP_WAIT, RESIGN,
 ADMIN_MENU, ATTEND_WAIT, POINTS_WAIT) = range(19)

T = {
"uz":{
"mm":"📋 Asosiy menyu:","su":"📝 Bir martalik Ro'yxatdan O'tish","bk":"🔙 Orqaga",
"nm":"👤 To'liq ismingizni kiriting (Familiya Ism):","sc":"🏫 Maktabingizni kiriting yoki tugmani bosing:",
"os":"🏫 Chartak TIM","cl":"📚 Sinfingizni tanlang:","ag":"🎂 Yoshingizni kiriting:",
"en":"🗣 Ingliz tili darajangizni tanlang:","fb":"✅ Obuna bo'ldim",
"fm":"📲 Ro'yxatni yakunlash uchun obuna bo'ling:\n\n📢 Telegram: {tg}\n📸 Instagram: {ig}\n\nObuna bo'lgach tugmani bosing.",
"nj":"⚠️ Siz hali @specialdebate ga qo'shilmagansiz!\nAvval obuna bo'ling, keyin tugmani bosing.",
"cv":"✅ <b>Ma'lumotlarni tasdiqlang:</b>\n\n👤 {name}\n🏫 {school}\n📚 {cls}\n🎂 {age}\n🗣 {eng}",
"ok":"✅ Tasdiqlash","ed":"✏️ Tahrirlash",
"ty":"🎉 Muvaffaqiyatli ro'yxatdan o'tdingiz!\nRahmat, {name}! 🎤",
"em":"Nimani tahrirlaysiz?","en1":"✏️ Ism","es":"🏫 Maktab","ec":"📚 Sinf","ea":"🎂 Yosh","ee":"🗣 Daraja",
"as":"⚠️ Allaqachon ro'yxatdan o'tgansiz! Qayta o'tmoqchimisiz?",
"ya":"✅ Ha","no":"❌ Yo'q","ns":"⚠️ Ro'yxatdan o'tmagansiz! Avval ro'yxatdan o'ting.",
"tb":"🔙 Orqaga","nv":"Yangi qiymatni kiriting:",
"ia":"⚠️ Raqam kiriting:","ie":"⚠️ Tugmalardan birini tanlang:",
},
"ru":{
"mm":"📋 Главное меню:","su":"📝 Единовременная Регистрация","bk":"🔙 Назад",
"nm":"👤 Введите полное имя (Фамилия Имя):","sc":"🏫 Введите школу или нажмите кнопку:",
"os":"🏫 Chartak TIM","cl":"📚 Выберите класс:","ag":"🎂 Введите возраст:",
"en":"🗣 Выберите уровень английского:","fb":"✅ Я подписался",
"fm":"📲 Подпишитесь на наши страницы:\n\n📢 Telegram: {tg}\n📸 Instagram: {ig}\n\nПосле подписки нажмите кнопку.",
"nj":"⚠️ Вы не подписались на @specialdebate!\nСначала подпишитесь, затем нажмите кнопку.",
"cv":"✅ <b>Подтвердите данные:</b>\n\n👤 {name}\n🏫 {school}\n📚 {cls}\n🎂 {age}\n🗣 {eng}",
"ok":"✅ Подтвердить","ed":"✏️ Редактировать",
"ty":"🎉 Регистрация успешна!\nСпасибо, {name}! 🎤",
"em":"Что изменить?","en1":"✏️ Имя","es":"🏫 Школа","ec":"📚 Класс","ea":"🎂 Возраст","ee":"🗣 Уровень",
"as":"⚠️ Вы уже зарегистрированы! Перерегистрироваться?",
"ya":"✅ Да","no":"❌ Нет","ns":"⚠️ Вы не зарегистрированы! Сначала зарегистрируйтесь.",
"tb":"🔙 Назад","nv":"Введите новое значение:",
"ia":"⚠️ Введите число:","ie":"⚠️ Выберите кнопку:",
},
"en":{
"mm":"📋 Main menu:","su":"📝 One-Time Registration","bk":"🔙 Back",
"nm":"👤 Enter your full name (Last First):","sc":"🏫 Type your school or press the button:",
"os":"🏫 Chartak TIM","cl":"📚 Select your class:","ag":"🎂 Enter your age:",
"en":"🗣 Select your English level:","fb":"✅ I followed",
"fm":"📲 Follow our pages to complete registration:\n\n📢 Telegram: {tg}\n📸 Instagram: {ig}\n\nAfter following, press the button.",
"nj":"⚠️ You haven't joined @specialdebate yet!\nPlease join first, then press the button.",
"cv":"✅ <b>Confirm your details:</b>\n\n👤 {name}\n🏫 {school}\n📚 {cls}\n🎂 {age}\n🗣 {eng}",
"ok":"✅ Confirm","ed":"✏️ Edit",
"ty":"🎉 Registration successful!\nThank you, {name}! Welcome to Special Debate! 🎤",
"em":"What to edit?","en1":"✏️ Name","es":"🏫 School","ec":"📚 Class","ea":"🎂 Age","ee":"🗣 Level",
"as":"⚠️ Already registered! Re-register?",
"ya":"✅ Yes","no":"❌ No","ns":"⚠️ You haven't signed up yet! Please sign up first.",
"tb":"🔙 Back","nv":"Enter new value:",
"ia":"⚠️ Please enter a number:","ie":"⚠️ Please press a button:",
},
}

def tr(lang, k): return T.get(lang, T["en"]).get(k, k)

def load_event():
    if os.path.exists(EVENT_FILE):
        with open(EVENT_FILE,"r",encoding="utf-8") as f: return json.load(f)
    return {"date":"TBA","topic":"TBA","photo":None}

def save_event(d):
    with open(EVENT_FILE,"w",encoding="utf-8") as f: json.dump(d,f,ensure_ascii=False)

def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE,"r") as f: return json.load(f)
    return []

def add_user(uid):
    users = load_users()
    if str(uid) not in users:
        users.append(str(uid))
        with open(USERS_FILE,"w") as f: json.dump(users,f)

def get_regs():
    if not os.path.exists(DATA_FILE): return {}
    wb = load_workbook(DATA_FILE); ws = wb.active; regs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]: regs[str(row[0])] = {"name":row[1],"school":row[2],"class":row[3],"age":row[4],"english":row[5],"date":row[6]}
    return regs

def save_reg(uid, d):
    wb = load_workbook(DATA_FILE) if os.path.exists(DATA_FILE) else openpyxl.Workbook()
    if not os.path.exists(DATA_FILE):
        ws = wb.active; ws.title="Registrations"
        ws.append(["UserID","Full Name","School","Class","Age","English","Registered At"])
    ws = wb.active
    todel = [r[0].row for r in ws.iter_rows(min_row=2) if str(r[0].value)==str(uid)]
    for r in reversed(todel): ws.delete_rows(r)
    ws.append([str(uid),d.get("name",""),d.get("school",""),d.get("class",""),d.get("age",""),d.get("english",""),datetime.now().strftime("%Y-%m-%d %H:%M")])
    wb.save(DATA_FILE)

def is_reg(uid): return str(uid) in get_regs()
def is_adm(uid): return uid in ADMIN_IDS

def lang_kb(): return ReplyKeyboardMarkup([["🇺🇿 O'zbek","🇷🇺 Русский","🇬🇧 English"]],resize_keyboard=True,one_time_keyboard=True)
def menu_kb(l): return ReplyKeyboardMarkup([[tr(l,"su")],[tr(l,"bk")]],resize_keyboard=True)
def class_kb(): return ReplyKeyboardMarkup([[f"{g}-01",f"{g}-02",f"{g}-03"] for g in range(5,12)],resize_keyboard=True,one_time_keyboard=True)

async def chk_member(bot, uid):
    try:
        m = await bot.get_chat_member(chat_id=TG_CHANNEL_ID, user_id=uid)
        return m.status in [ChatMember.MEMBER, ChatMember.ADMINISTRATOR, ChatMember.OWNER]
    except: return True

async def start(u: Update, c: ContextTypes.DEFAULT_TYPE):
    c.user_data.clear(); add_user(u.effective_user.id)
    await u.message.reply_text("👋 Welcome to <b>Special Debate</b>!\n\nTilni tanlang / Choose language / Выберите язык:",parse_mode="HTML",reply_markup=lang_kb())
    return LANG

async def sel_lang(u: Update, c: ContextTypes.DEFAULT_TYPE):
    t = u.message.text
    l = "uz" if "O'zbek" in t else "ru" if "Русский" in t else "en"
    c.user_data["lang"] = l
    await u.message.reply_text(tr(l,"mm"), reply_markup=menu_kb(l))
    return MAIN_MENU

async def main_menu(u: Update, c: ContextTypes.DEFAULT_TYPE):
    l = c.user_data.get("lang","en"); t = u.message.text
    if t == tr(l,"bk"):
        await u.message.reply_text("Tilni tanlang / Choose language / Выберите язык:",reply_markup=lang_kb()); return LANG
    if t == tr(l,"su"):
        if is_reg(u.effective_user.id):
            await u.message.reply_text(tr(l,"as"),reply_markup=ReplyKeyboardMarkup([[tr(l,"ya"),tr(l,"no")]],resize_keyboard=True)); return RESIGN
        c.user_data["s"] = {}
        await u.message.reply_text(tr(l,"nm"),reply_markup=ReplyKeyboardMarkup([[tr(l,"bk")]],resize_keyboard=True)); return SU_NAME
    return MAIN_MENU

async def resign(u: Update, c: ContextTypes.DEFAULT_TYPE):
    l = c.user_data.get("lang","en")
    if u.message.text == tr(l,"ya"):
        c.user_data["s"] = {}
        await u.message.reply_text(tr(l,"nm"),reply_markup=ReplyKeyboardMarkup([[tr(l,"bk")]],resize_keyboard=True)); return SU_NAME
    await u.message.reply_text(tr(l,"mm"),reply_markup=menu_kb(l)); return MAIN_MENU

async def su_name(u: Update, c: ContextTypes.DEFAULT_TYPE):
    l = c.user_data.get("lang","en"); t = u.message.text
    if t == tr(l,"bk"): await u.message.reply_text(tr(l,"mm"),reply_markup=menu_kb(l)); return MAIN_MENU
    c.user_data["s"]["name"] = t
    await u.message.reply_text(tr(l,"sc"),reply_markup=ReplyKeyboardMarkup([[tr(l,"os")],[tr(l,"bk")]],resize_keyboard=True)); return SU_SCHOOL

async def su_school(u: Update, c: ContextTypes.DEFAULT_TYPE):
    l = c.user_data.get("lang","en"); t = u.message.text
    if t == tr(l,"bk"): await u.message.reply_text(tr(l,"nm"),reply_markup=ReplyKeyboardMarkup([[tr(l,"bk")]],resize_keyboard=True)); return SU_NAME
    chartak = "Chartak TIM" in t
    c.user_data["s"]["school"] = "Chartak TIM" if chartak else t
    c.user_data["s"]["chartak"] = chartak
    if chartak: await u.message.reply_text(tr(l,"cl"),reply_markup=class_kb()); return SU_CLASS
    c.user_data["s"]["class"] = "-"
    await u.message.reply_text(tr(l,"ag"),reply_markup=ReplyKeyboardMarkup([[tr(l,"bk")]],resize_keyboard=True)); return SU_AGE

async def su_class(u: Update, c: ContextTypes.DEFAULT_TYPE):
    l = c.user_data.get("lang","en"); t = u.message.text
    if t == tr(l,"bk"): await u.message.reply_text(tr(l,"sc"),reply_markup=ReplyKeyboardMarkup([[tr(l,"os")],[tr(l,"bk")]],resize_keyboard=True)); return SU_SCHOOL
    c.user_data["s"]["class"] = t
    await u.message.reply_text(tr(l,"ag"),reply_markup=ReplyKeyboardMarkup([[tr(l,"bk")]],resize_keyboard=True)); return SU_AGE

async def su_age(u: Update, c: ContextTypes.DEFAULT_TYPE):
    l = c.user_data.get("lang","en"); t = u.message.text
    if t == tr(l,"bk"):
        if c.user_data["s"].get("chartak"): await u.message.reply_text(tr(l,"cl"),reply_markup=class_kb()); return SU_CLASS
        await u.message.reply_text(tr(l,"sc"),reply_markup=ReplyKeyboardMarkup([[tr(l,"os")],[tr(l,"bk")]],resize_keyboard=True)); return SU_SCHOOL
    if not t.isdigit(): await u.message.reply_text(tr(l,"ia")); return SU_AGE
    c.user_data["s"]["age"] = t
    await u.message.reply_text(tr(l,"en"),reply_markup=ReplyKeyboardMarkup([["A1","A2"],["B1","B2"],["C1","C2"],[tr(l,"bk")]],resize_keyboard=True)); return SU_ENGLISH

async def su_english(u: Update, c: ContextTypes.DEFAULT_TYPE):
    l = c.user_data.get("lang","en"); t = u.message.text.strip().upper()
    if t in [tr(l,"bk").upper(),"🔙 ORQAGA","🔙 НАЗАД","🔙 BACK"]:
        await u.message.reply_text(tr(l,"ag"),reply_markup=ReplyKeyboardMarkup([[tr(l,"bk")]],resize_keyboard=True)); return SU_AGE
    if t not in ["A1","A2","B1","B2","C1","C2"]: await u.message.reply_text(tr(l,"ie")); return SU_ENGLISH
    c.user_data["s"]["english"] = t
    await u.message.reply_text(tr(l,"fm").format(tg=TG_CHANNEL,ig=IG_PAGE),parse_mode="HTML",reply_markup=ReplyKeyboardMarkup([[tr(l,"fb")]],resize_keyboard=True)); return SU_FOLLOW

async def su_follow(u: Update, c: ContextTypes.DEFAULT_TYPE):
    l = c.user_data.get("lang","en")
    if not await chk_member(u.get_bot(), u.effective_user.id):
        await u.message.reply_text(tr(l,"nj"),reply_markup=ReplyKeyboardMarkup([[tr(l,"fb")]],resize_keyboard=True)); return SU_FOLLOW
    s = c.user_data["s"]
    await u.message.reply_text(tr(l,"cv").format(name=s.get("name",""),school=s.get("school",""),cls=s.get("class","-"),age=s.get("age",""),eng=s.get("english","")),
        parse_mode="HTML",reply_markup=ReplyKeyboardMarkup([[tr(l,"ok"),tr(l,"ed")]],resize_keyboard=True)); return SU_CONFIRM

async def su_confirm(u: Update, c: ContextTypes.DEFAULT_TYPE):
    l = c.user_data.get("lang","en"); t = u.message.text; cid = u.effective_chat.id
    if t == tr(l,"ok"):
        save_reg(u.effective_user.id, c.user_data["s"])
        thanks = await u.message.reply_text(tr(l,"ty").format(name=c.user_data["s"].get("name","")),reply_markup=ReplyKeyboardRemove())
        mid = u.message.message_id
        for i in range(mid, max(mid-80,0), -1):
            if i == thanks.message_id: continue
            try: await c.bot.delete_message(chat_id=cid, message_id=i)
            except: pass
        await c.bot.send_message(chat_id=cid, text=tr(l,"mm"), reply_markup=menu_kb(l)); return MAIN_MENU
    if t == tr(l,"ed"):
        s = c.user_data["s"]
        opts = [[tr(l,"en1"),tr(l,"es")]]
        if s.get("chartak"): opts.append([tr(l,"ec")])
        opts.append([tr(l,"ea"),tr(l,"ee")])
        await u.message.reply_text(tr(l,"em"),reply_markup=ReplyKeyboardMarkup(opts,resize_keyboard=True)); return SU_EDIT_MENU
    return SU_CONFIRM

async def su_edit_menu(u: Update, c: ContextTypes.DEFAULT_TYPE):
    l = c.user_data.get("lang","en"); c.user_data["ef"] = u.message.text
    await u.message.reply_text(tr(l,"nv"),reply_markup=ReplyKeyboardMarkup([[tr(l,"bk")]],resize_keyboard=True)); return SU_EDIT_FIELD

async def su_edit_field(u: Update, c: ContextTypes.DEFAULT_TYPE):
    l = c.user_data.get("lang","en"); t = u.message.text; s = c.user_data["s"]
    if t == tr(l,"bk"):
        await u.message.reply_text(tr(l,"cv").format(name=s.get("name",""),school=s.get("school",""),cls=s.get("class","-"),age=s.get("age",""),eng=s.get("english","")),
            parse_mode="HTML",reply_markup=ReplyKeyboardMarkup([[tr(l,"ok"),tr(l,"ed")]],resize_keyboard=True)); return SU_CONFIRM
    f = c.user_data.get("ef","")
    if f==tr(l,"en1"): s["name"]=t
    elif f==tr(l,"es"): s["school"]=t
    elif f==tr(l,"ec"): s["class"]=t
    elif f==tr(l,"ea"): s["age"]=t
    elif f==tr(l,"ee"): s["english"]=t.upper()
    await u.message.reply_text(tr(l,"cv").format(name=s.get("name",""),school=s.get("school",""),cls=s.get("class","-"),age=s.get("age",""),eng=s.get("english","")),
        parse_mode="HTML",reply_markup=ReplyKeyboardMarkup([[tr(l,"ok"),tr(l,"ed")]],resize_keyboard=True)); return SU_CONFIRM

ATTEND_FILE = "attendance.json"
POINTS_FILE = "points.json"

def load_attendance():
    if os.path.exists(ATTEND_FILE):
        with open(ATTEND_FILE,"r",encoding="utf-8") as f: return json.load(f)
    return {}

def save_attendance(d):
    with open(ATTEND_FILE,"w",encoding="utf-8") as f: json.dump(d,f,ensure_ascii=False)

def load_points():
    if os.path.exists(POINTS_FILE):
        with open(POINTS_FILE,"r",encoding="utf-8") as f: return json.load(f)
    return {}

def save_points(d):
    with open(POINTS_FILE,"w",encoding="utf-8") as f: json.dump(d,f,ensure_ascii=False)

def admin_main_kb():
    return ReplyKeyboardMarkup([["✅ Check Attendance","🏆 Announce Points"],["🔙 Back"]],resize_keyboard=True)

async def cmd_admin_panel(u: Update, c: ContextTypes.DEFAULT_TYPE):
    if not is_adm(u.effective_user.id): return ConversationHandler.END
    c.user_data["attend_checked"] = {}
    await u.message.reply_text("🛠 <b>Admin Panel</b>\n\nChoose an action:",parse_mode="HTML",reply_markup=admin_main_kb())
    return ADMIN_MENU

async def admin_menu(u: Update, c: ContextTypes.DEFAULT_TYPE):
    t = u.message.text
    if t == "🔙 Back":
        await u.message.reply_text("Exited admin panel.",reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END
    if t == "✅ Check Attendance":
        regs = get_regs()
        if not regs:
            await u.message.reply_text("⚠️ No registered students yet.",reply_markup=admin_main_kb())
            return ADMIN_MENU
        c.user_data["attend_checked"] = {}
        c.user_data["attend_names"] = {uid: d["name"] for uid,d in regs.items()}
        await _send_attendance_list(u, c)
        return ATTEND_WAIT
    if t == "🏆 Announce Points":
        points = load_points()
        if not points:
            await u.message.reply_text("📊 No points recorded yet.",reply_markup=admin_main_kb())
            return ADMIN_MENU
        lines = ["🏆 <b>Points Leaderboard</b>\n"]
        sorted_pts = sorted(points.items(), key=lambda x: x[1]["hours"], reverse=True)
        for i,(uid,d) in enumerate(sorted_pts,1):
            lines.append(f"{i}. {d['name']} — {d['hours']:.1f} hrs ({d['sessions']} sessions)")
        await u.message.reply_text("\n".join(lines),parse_mode="HTML",reply_markup=admin_main_kb())
        return ADMIN_MENU
    return ADMIN_MENU

async def _send_attendance_list(u: Update, c: ContextTypes.DEFAULT_TYPE):
    names = c.user_data.get("attend_names", {})
    checked = c.user_data.get("attend_checked", {})
    rows = []
    for uid, name in names.items():
        tick = "✅ " if uid in checked else "⬜ "
        rows.append([tick + name])
    rows.append(["✔️ Confirm Attendance"])
    rows.append(["🔙 Back to Menu"])
    await u.message.reply_text(
        "👥 <b>Select present students</b>\nTap a name to mark ✅ / unmark ⬜\nThen press <b>Confirm Attendance</b>",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(rows, resize_keyboard=True)
    )

async def attend_handler(u: Update, c: ContextTypes.DEFAULT_TYPE):
    t = u.message.text
    if t == "🔙 Back to Menu":
        c.user_data["attend_checked"] = {}
        await u.message.reply_text("🛠 Admin Panel:",reply_markup=admin_main_kb())
        return ADMIN_MENU
    if t == "✔️ Confirm Attendance":
        checked = c.user_data.get("attend_checked", {})
        names = c.user_data.get("attend_names", {})
        if not checked:
            await u.message.reply_text("⚠️ No students marked yet. Tap names to mark attendance first.")
            await _send_attendance_list(u, c)
            return ATTEND_WAIT
        today = datetime.now().strftime("%Y-%m-%d")
        attendance = load_attendance()
        if today not in attendance: attendance[today] = []
        for uid in checked:
            if uid not in attendance[today]: attendance[today].append(uid)
        save_attendance(attendance)
        points = load_points()
        for uid, name in checked.items():
            if uid not in points: points[uid] = {"name": name, "hours": 0.0, "sessions": 0}
            points[uid]["hours"] = round(points[uid]["hours"] + 1.5, 1)
            points[uid]["sessions"] += 1
            points[uid]["name"] = name
        save_points(points)
        lines = [f"✅ <b>Attendance confirmed for {today}</b>\n"]
        lines.append(f"📋 {len(checked)} students marked present:\n")
        for uid, name in checked.items():
            lines.append(f"• {name} → +1.5 hrs (total: {points[uid]['hours']:.1f} hrs)")
        await u.message.reply_text("\n".join(lines),parse_mode="HTML",reply_markup=admin_main_kb())
        c.user_data["attend_checked"] = {}
        return ADMIN_MENU
    names = c.user_data.get("attend_names", {})
    checked = c.user_data.get("attend_checked", {})
    clean = t.replace("✅ ","").replace("⬜ ","").strip()
    matched_uid = None
    for uid, name in names.items():
        if name == clean:
            matched_uid = uid; break
    if matched_uid:
        if matched_uid in checked: del checked[matched_uid]
        else: checked[matched_uid] = names[matched_uid]
        c.user_data["attend_checked"] = checked
    await _send_attendance_list(u, c)
    return ATTEND_WAIT

async def cmd_getinfo(u: Update, c: ContextTypes.DEFAULT_TYPE):
    if not is_adm(u.effective_user.id): return
    regs = get_regs()
    if not regs: await u.message.reply_text("📭 No registrations yet."); return
    lines = [f"📋 <b>Registrations ({len(regs)})</b>\n"]
    for uid,d in regs.items(): lines.append(f"👤 {d['name']} | 🏫 {d['school']} | 📚 {d.get('class','-')} | 🎂 {d['age']} | 🗣 {d['english']} | 📅 {d['date']}")
    text = "\n".join(lines)
    for i in range(0,len(text),4000): await u.message.reply_text(text[i:i+4000],parse_mode="HTML")

async def cmd_broadcast(u: Update, c: ContextTypes.DEFAULT_TYPE):
    if not is_adm(u.effective_user.id): return ConversationHandler.END
    users = load_users()
    await u.message.reply_text(f"📢 Send your message to broadcast to <b>{len(users)}</b> users.\n(Text, photo or video)",parse_mode="HTML")
    return BC_WAIT

async def bc_send(u: Update, c: ContextTypes.DEFAULT_TYPE):
    if not is_adm(u.effective_user.id): return ConversationHandler.END
    users = load_users(); ok = 0; fail = 0
    for uid in users:
        try:
            if u.message.photo: await c.bot.send_photo(int(uid),photo=u.message.photo[-1].file_id,caption=u.message.caption or "")
            elif u.message.video: await c.bot.send_video(int(uid),video=u.message.video.file_id,caption=u.message.caption or "")
            else: await c.bot.send_message(int(uid),text=u.message.text)
            ok+=1
        except: fail+=1
    await u.message.reply_text(f"✅ Broadcast done! Sent: {ok} | Failed: {fail}")
    return ConversationHandler.END

async def cmd_setdate(u: Update, c: ContextTypes.DEFAULT_TYPE):
    if not is_adm(u.effective_user.id): return ConversationHandler.END
    await u.message.reply_text("📅 Enter the new event date:\nExample: 25th of March 2025")
    return SD_WAIT

async def sd_save(u: Update, c: ContextTypes.DEFAULT_TYPE):
    ev = load_event(); ev["date"] = f"📅 The next Debate is on {u.message.text}"; save_event(ev)
    await u.message.reply_text(f"✅ Date updated!\n{ev['date']}")
    return ConversationHandler.END

async def cmd_settopic(u: Update, c: ContextTypes.DEFAULT_TYPE):
    if not is_adm(u.effective_user.id): return ConversationHandler.END
    await u.message.reply_text("📌 Enter the debate topic:")
    return ST_WAIT

async def st_save(u: Update, c: ContextTypes.DEFAULT_TYPE):
    ev = load_event(); ev["topic"] = f"📌 The topic is: {u.message.text}"; save_event(ev)
    await u.message.reply_text(f"✅ Topic updated!\n{ev['topic']}")
    return ConversationHandler.END

async def cmd_setphoto(u: Update, c: ContextTypes.DEFAULT_TYPE):
    if not is_adm(u.effective_user.id): return ConversationHandler.END
    await u.message.reply_text("🖼 Send the new event poster photo:")
    return SP_WAIT

async def sp_save(u: Update, c: ContextTypes.DEFAULT_TYPE):
    if not u.message.photo: await u.message.reply_text("⚠️ Please send a photo."); return SP_WAIT
    ev = load_event(); ev["photo"] = u.message.photo[-1].file_id; save_event(ev)
    await u.message.reply_text("✅ Photo updated!"); return ConversationHandler.END

async def cmd_admin(u: Update, c: ContextTypes.DEFAULT_TYPE):
    if not is_adm(u.effective_user.id): return
    await u.message.reply_text("🛠 <b>Admin Commands</b>\n\n/panel — Open admin panel (attendance + points)\n/getinfo — All registrations\n/broadcast — Message all users\n/setdate — Change event date\n/settopic — Change topic\n/setphoto — Change poster",parse_mode="HTML")

def main():
    app = Application.builder().token(BOT_TOKEN).build()

    bc_conv  = ConversationHandler(entry_points=[CommandHandler("broadcast",cmd_broadcast)],
        states={BC_WAIT:[MessageHandler(filters.TEXT&~filters.COMMAND,bc_send),MessageHandler(filters.PHOTO,bc_send),MessageHandler(filters.VIDEO,bc_send)]},
        fallbacks=[CommandHandler("start",start)])
    sd_conv  = ConversationHandler(entry_points=[CommandHandler("setdate",cmd_setdate)],
        states={SD_WAIT:[MessageHandler(filters.TEXT&~filters.COMMAND,sd_save)]},
        fallbacks=[CommandHandler("start",start)])
    st_conv  = ConversationHandler(entry_points=[CommandHandler("settopic",cmd_settopic)],
        states={ST_WAIT:[MessageHandler(filters.TEXT&~filters.COMMAND,st_save)]},
        fallbacks=[CommandHandler("start",start)])
    sp_conv  = ConversationHandler(entry_points=[CommandHandler("setphoto",cmd_setphoto)],
        states={SP_WAIT:[MessageHandler(filters.PHOTO,sp_save)]},
        fallbacks=[CommandHandler("start",start)])

    admin_conv = ConversationHandler(entry_points=[CommandHandler("panel",cmd_admin_panel)],
        states={
            ADMIN_MENU:[MessageHandler(filters.TEXT&~filters.COMMAND,admin_menu)],
            ATTEND_WAIT:[MessageHandler(filters.TEXT&~filters.COMMAND,attend_handler)],
        },
        fallbacks=[CommandHandler("start",start)], allow_reentry=True)

    main_conv = ConversationHandler(entry_points=[CommandHandler("start",start)],
        states={
            LANG:[MessageHandler(filters.TEXT&~filters.COMMAND,sel_lang)],
            MAIN_MENU:[MessageHandler(filters.TEXT&~filters.COMMAND,main_menu)],
            RESIGN:[MessageHandler(filters.TEXT&~filters.COMMAND,resign)],
            SU_NAME:[MessageHandler(filters.TEXT&~filters.COMMAND,su_name)],
            SU_SCHOOL:[MessageHandler(filters.TEXT&~filters.COMMAND,su_school)],
            SU_CLASS:[MessageHandler(filters.TEXT&~filters.COMMAND,su_class)],
            SU_AGE:[MessageHandler(filters.TEXT&~filters.COMMAND,su_age)],
            SU_ENGLISH:[MessageHandler(filters.TEXT&~filters.COMMAND,su_english)],
            SU_FOLLOW:[MessageHandler(filters.TEXT&~filters.COMMAND,su_follow)],
            SU_CONFIRM:[MessageHandler(filters.TEXT&~filters.COMMAND,su_confirm)],
            SU_EDIT_MENU:[MessageHandler(filters.TEXT&~filters.COMMAND,su_edit_menu)],
            SU_EDIT_FIELD:[MessageHandler(filters.TEXT&~filters.COMMAND,su_edit_field)],
        },
        fallbacks=[CommandHandler("start",start)], allow_reentry=True)

    app.add_handler(bc_conv)
    app.add_handler(sd_conv)
    app.add_handler(st_conv)
    app.add_handler(sp_conv)
    app.add_handler(admin_conv)
    app.add_handler(main_conv)
    app.add_handler(CommandHandler("getinfo",cmd_getinfo))
    app.add_handler(CommandHandler("admin",cmd_admin))

    print("✅ Special Debate Bot is running...")
    app.run_polling()

if __name__ == "__main__":
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    main()  
